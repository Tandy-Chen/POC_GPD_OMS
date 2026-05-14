"""Pipeline orchestration."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .config import (
    DEFAULT_INPUT_DIR,
    DEFAULT_OUTPUT_DIR,
    DEFAULT_SOURCE_SHEET,
    OUTPUT_FILE_EXTENSION,
    build_output_file_prefix,
)
from .reader import ValidationSummary, read_source_rows, validate_source
from .reports.exception_report import render_exception_report
from .reports.matrix_report import render_matrix_report
from .reports.performance_report import render_performance_report
from .reports.style import apply_body_borders, apply_header_style, apply_standard_layout, auto_fit_columns
from .reports.summary_report import render_summary_report


def _build_snapshot_sheet_name(source_sheet_name: str) -> str:
    # Excel sheet name max length is 31.
    return f"來源資料_{source_sheet_name}"[:31]


def _unique_sheet_title(workbook: Workbook, preferred_title: str) -> str:
    if preferred_title not in workbook.sheetnames:
        return preferred_title

    base = preferred_title[:28]
    index = 2
    while True:
        candidate = f"{base}_{index}"
        if candidate not in workbook.sheetnames:
            return candidate
        index += 1


def append_source_sheet_snapshot(
    workbook: Workbook,
    source_path: Path,
    source_sheet_name: str,
) -> str:
    snapshot_wb = load_workbook(source_path, read_only=True, data_only=True)
    try:
        source_ws = snapshot_wb[source_sheet_name]
        preferred_title = _build_snapshot_sheet_name(source_sheet_name)
        snapshot_title = _unique_sheet_title(workbook, preferred_title)
        target_ws = workbook.create_sheet(snapshot_title)

        for row in source_ws.iter_rows(values_only=True):
            target_ws.append(list(row))

        content_end_row = target_ws.max_row
        apply_body_borders(target_ws, start_row=1, end_row=content_end_row)
        if target_ws.max_row >= 1:
            apply_header_style(target_ws, 1)
            apply_standard_layout(target_ws, "A2")
        auto_fit_columns(target_ws, max_row=content_end_row)
    finally:
        snapshot_wb.close()

    return snapshot_title


def run_validate(
    input_path: str | Path | None = None,
    sheet_name: str = DEFAULT_SOURCE_SHEET,
    input_dir: str | Path = DEFAULT_INPUT_DIR,
) -> ValidationSummary:
    return validate_source(input_path=input_path, sheet_name=sheet_name, input_dir=input_dir)


def build_output_filename(now: datetime | None = None) -> str:
    current_time = now or datetime.now()
    ts = current_time.strftime("%Y%m%d_%H%M")
    prefix = build_output_file_prefix(current_time)
    return f"{prefix}_{ts}.{OUTPUT_FILE_EXTENSION}"


def run_process(
    input_path: str | Path | None = None,
    sheet_name: str = DEFAULT_SOURCE_SHEET,
    input_dir: str | Path = DEFAULT_INPUT_DIR,
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
    now: datetime | None = None,
) -> dict:
    """Run full processing pipeline (v0.3+)."""
    rows, actual_source_path = read_source_rows(input_path=input_path, sheet_name=sheet_name, input_dir=input_dir)
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True, parents=True)
    output_file = output_path / build_output_filename(now=now)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    render_matrix_report(workbook, rows, actual_source_path.name, now=now)
    render_exception_report(workbook, rows, now=now)
    render_performance_report(workbook, rows, now=now)
    source_snapshot_sheet_name = append_source_sheet_snapshot(
        workbook=workbook,
        source_path=actual_source_path,
        source_sheet_name=sheet_name,
    )
    render_summary_report(workbook, source_path=actual_source_path, source_rows=len(rows), now=now)
    workbook.save(output_file)

    return {
        "source_path": actual_source_path,
        "source_snapshot_sheet": source_snapshot_sheet_name,
        "output_dir": output_path,
        "output_file": output_file,
        "rows_processed": len(rows),
        "sheet_names": workbook.sheetnames,
    }

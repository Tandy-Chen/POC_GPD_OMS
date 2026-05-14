"""Styling helpers."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
import unicodedata

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

HEADER_COLOR = "003399"
TAIPEI_TZ = timezone(timedelta(hours=8))


def apply_header_style(ws: Worksheet, row_index: int = 1) -> None:
    for cell in ws[row_index]:
        cell.font = Font(bold=True, color=HEADER_COLOR)


def apply_body_borders(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int = 1,
    end_col: int | None = None,
) -> None:
    if end_row < start_row:
        return
    max_col = end_col or ws.max_column
    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).border = border


def append_merged_note(ws: Worksheet, text: str, now_row: bool = True) -> int:
    if now_row:
        ws.append([])
    ws.append([text])
    row_idx = ws.max_row
    if ws.max_column > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=ws.max_column)
    note_cell = ws.cell(row=row_idx, column=1)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    return row_idx


def _display_width(value: object) -> int:
    text = str(value or "")
    width = 0
    for ch in text:
        width += 2 if unicodedata.east_asian_width(ch) in {"F", "W"} else 1
    return width


def auto_fit_columns(
    ws: Worksheet,
    min_width: int = 12,
    explicit_widths: dict[str, int] | None = None,
    max_row: int | None = None,
) -> None:
    widths = explicit_widths or {}
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        bounded_cells = column_cells if max_row is None else [c for c in column_cells if c.row <= max_row]
        if not bounded_cells:
            max_length = 0
        else:
            max_length = max(_display_width(cell.value) for cell in bounded_cells)
        ws.column_dimensions[column_letter].width = max(widths.get(column_letter, 0), max(min_width, max_length + 2))


def apply_standard_layout(ws: Worksheet, freeze_panes: str = "A2") -> None:
    ws.freeze_panes = freeze_panes
    ws.sheet_view.showGridLines = True


def append_stat_date(ws: Worksheet, now: datetime | None = None) -> None:
    stat_time = now.astimezone(TAIPEI_TZ) if now else datetime.now(TAIPEI_TZ)
    append_merged_note(ws, f"統計日期：{stat_time.strftime('%Y-%m-%d %H:%M')} UTC+8")

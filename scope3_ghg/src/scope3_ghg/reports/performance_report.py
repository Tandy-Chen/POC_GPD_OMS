"""Performance report renderer."""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook

from ..aggregators import group_rows_by_department
from ..config import DEPT_ORDER
from ..kpi import compute_performance
from ..models import ContractRow
from .style import append_stat_date, apply_body_borders, apply_header_style, apply_standard_layout, auto_fit_columns

MAIN_SHEET_NAME = "碳盤查績效統計（資分）"


def _write_performance_sheet(ws, title: str, rows: list[ContractRow], now: datetime | None = None) -> None:
    perf = compute_performance(rows)
    ws.append(["1.3範疇三供應鏈碳管理：", "KPI計算", f"{title.replace('碳盤查績效統計_', '').replace(MAIN_SHEET_NAME, '資分')}實績現況"])
    ws.append(["1.1 盤查件數達成率", perf.kpiA1_pct, perf.scoreA1])
    ws.append(["1.2 盤查契約金額達成率", perf.kpiA2_pct, perf.scoreA2])
    ws.append(["1.3 碳足跡條款落實達成率", perf.kpiA3_pct, perf.scoreA3])
    ws.append(["2 供應商淨零承諾達成率", perf.kpiB_ratio, perf.scoreB])
    ws.append(["3 ISO 14064-1輔導證書達成率", perf.kpiC_ratio, perf.scoreC])
    ws.append([f"{title.replace('碳盤查績效統計_', '').replace(MAIN_SHEET_NAME, '資分')}累計得分試算", perf.total_calc, perf.total_raw])
    content_end_row = ws.max_row
    apply_header_style(ws, 1)
    apply_body_borders(ws, start_row=1, end_row=content_end_row)
    append_stat_date(ws, now=now)
    apply_standard_layout(ws, "A2")
    auto_fit_columns(ws, explicit_widths={"A": 34, "B": 14, "C": 16}, max_row=content_end_row)


def render_performance_report(workbook: Workbook, rows: list[ContractRow], now: datetime | None = None) -> None:
    grouped = group_rows_by_department(rows)

    main_ws = workbook.create_sheet(MAIN_SHEET_NAME)
    _write_performance_sheet(main_ws, MAIN_SHEET_NAME, rows, now=now)

    for dept in DEPT_ORDER:
        ws = workbook.create_sheet(f"碳盤查績效統計_{dept}")
        _write_performance_sheet(ws, f"碳盤查績效統計_{dept}", grouped.get(dept, []), now=now)

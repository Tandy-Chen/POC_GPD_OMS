"""Source Excel read and validation."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .config import (
    DEFAULT_INPUT_DIR,
    DEFAULT_SOURCE_SHEET,
    REQUIRED_FIELDS,
    SOURCE_COLUMN_ALIASES,
    SOURCE_FILE_PATTERN_EXT,
    SOURCE_FILE_PATTERN_NAME,
)
from .models import ContractRow
from .normalize import normalize_compact_text, normalize_header_name, to_safe_number


class SourceReadError(RuntimeError):
    """Raised when source file cannot be read or validated."""


def locate_source_file(input_dir: str | Path = DEFAULT_INPUT_DIR) -> Path:
    """Find source file by pattern matching: .xlsx + 範疇三業務管理報表."""
    dir_path = Path(input_dir)
    if not dir_path.is_dir():
        raise SourceReadError(f"Input directory not found: {dir_path}")

    candidates = [
        p
        for p in dir_path.glob(f"*.{SOURCE_FILE_PATTERN_EXT}")
        if SOURCE_FILE_PATTERN_NAME in p.name
    ]
    if not candidates:
        raise SourceReadError(
            f"No source file found in {dir_path} matching *.{SOURCE_FILE_PATTERN_EXT} + '{SOURCE_FILE_PATTERN_NAME}'"
        )
    if len(candidates) > 1:
        raise SourceReadError(
            f"Multiple source files found in {dir_path}: {[c.name for c in candidates]}. Please keep only one."
        )
    return candidates[0]


@dataclass(slots=True)
class ValidationSummary:
    source_rows: int
    valid_rows: int


def _build_header_index(columns: list[str]) -> dict[str, int]:
    normalized_headers = [normalize_header_name(col) for col in columns]
    index_map: dict[str, int] = {}

    for field_key, aliases in SOURCE_COLUMN_ALIASES.items():
        hit = -1
        for alias in aliases:
            try:
                hit = normalized_headers.index(normalize_header_name(alias))
                break
            except ValueError:
                continue
        if hit >= 0:
            index_map[field_key] = hit

    missing = [f for f in REQUIRED_FIELDS if f not in index_map]
    if missing:
        detail = ", ".join(
            f"{field}=>{' / '.join(SOURCE_COLUMN_ALIASES[field])}" for field in missing
        )
        raise SourceReadError(f"Missing required columns: {detail}")
    return index_map


def read_source_rows(
    input_path: str | Path | None = None,
    sheet_name: str = DEFAULT_SOURCE_SHEET,
    input_dir: str | Path = DEFAULT_INPUT_DIR,
) -> tuple[list[ContractRow], Path]:
    """Read source rows and return (rows, actual_source_path)."""
    if input_path is None:
        actual_path = locate_source_file(input_dir=input_dir)
    else:
        actual_path = Path(input_path)

    if not actual_path.exists():
        raise SourceReadError(f"Input file not found: {actual_path}")

    try:
        wb = load_workbook(actual_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise SourceReadError(f"Unable to read Excel: {exc}") from exc

    if sheet_name not in wb.sheetnames:
        raise SourceReadError(f"Sheet not found: {sheet_name}")

    ws = wb[sheet_name]
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise SourceReadError("Source sheet is empty") from exc

    header_index = _build_header_index([str(c or "") for c in header_row])

    rows: list[ContractRow] = []
    for row in iterator:
        row_data = list(row)

        def pick(field: str) -> object:
            idx = header_index[field]
            if idx >= len(row_data):
                return ""
            return row_data[idx]

        rows.append(
            ContractRow(
                contract_id=str(pick("contract_id") or ""),
                contract_name=str(pick("contract_name") or ""),
                contract_amount=to_safe_number(pick("contract_amount")),
                requester_unit=str(pick("requester_unit") or ""),
                requester_name=str(pick("requester_name") or ""),
                contract_status_raw=str(pick("contract_status_raw") or ""),
                kpi_key=normalize_compact_text(pick("kpi_key")),
                vendor_name_raw=str(pick("vendor_name_raw") or ""),
                net_zero_commitment_raw=str(pick("net_zero_commitment_raw") or ""),
                iso14064_evidence_raw=str(
                    pick("iso14064_evidence_raw") if "iso14064_evidence_raw" in header_index else ""
                ),
            )
        )

    wb.close()
    return rows, actual_path


def validate_source(
    input_path: str | Path | None = None,
    sheet_name: str = DEFAULT_SOURCE_SHEET,
    input_dir: str | Path = DEFAULT_INPUT_DIR,
) -> ValidationSummary:
    rows, actual_path = read_source_rows(input_path=input_path, sheet_name=sheet_name, input_dir=input_dir)
    return ValidationSummary(source_rows=len(rows), valid_rows=len(rows))

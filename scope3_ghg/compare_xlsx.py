import openpyxl
import os

def get_sheet_data(wb, sheet_name, rows_limit=None, skip_empty=False, cols_indices=None):
    if sheet_name not in wb.sheetnames:
        # Try to find a partial match for "碳盤查績效統計"
        matches = [s for s in wb.sheetnames if "碳盤查績效統計" in s]
        if matches:
            sheet_name = matches[0]
        else:
            return None
    
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        if skip_empty and not any(row):
            continue
        
        if cols_indices:
            row_data = [row[i] if i < len(row) else None for i in cols_indices]
        else:
            row_data = list(row)
            
        data.append(row_data)
        if rows_limit and len(data) >= rows_limit:
            break
    return data

def process_matrix_sheet(wb):
    sheet_name = "統計分析_矩陣"
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    max_col = ws.max_column
    # Indices for first two and last two columns
    indices = [0, 1, max_col - 2, max_col - 1]
    data = []
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        row_data = [row[i] if i < len(row) else None for i in indices]
        data.append(row_data)
    return data

old_path = r"C:\Tandy資料夾\APPdata\supply-chain-carbon-management\SCOPE3_Report_GE115_各單位績效_20260511_0843.xlsx"
new_path = r"C:\Tandy資料夾\APPdata\supply-chain-carbon-management\scope3_ghg\outputs\SCOPE3_Report_GE115_各單位績效_20260513_1504.xlsx"

print("--- Comparison Report ---")

for label, path in [("OLD", old_path), ("NEW", new_path)]:
    print(f"\n[{label}] {os.path.basename(path)}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        
        # A. 統計分析_矩陣
        print(f"--- 統計分析_矩陣 (前兩欄 & 最後兩欄) ---")
        matrix_data = process_matrix_sheet(wb)
        if matrix_data:
            for row in matrix_data:
                print(row)
        else:
            print("Sheet '統計分析_矩陣' not found.")
            
        # B. 碳盤查績效統計
        print(f"--- 碳盤查績效統計 (前 7 列, 前 3 欄) ---")
        perf_data = get_sheet_data(wb, "碳盤查績效統計", rows_limit=7, cols_indices=[0, 1, 2])
        if perf_data:
            for row in perf_data:
                print(row)
        else:
            print("Sheet '碳盤查績效統計' or similar not found.")
            
    except Exception as e:
        print(f"Error processing {label}: {e}")


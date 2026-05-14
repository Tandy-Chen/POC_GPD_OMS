import openpyxl
import os

files = [
    (r"c:\Tandy資料夾\APPdata\supply-chain-carbon-management\SCOPE3_Report_GE115_各單位績效_20260511_0843.xlsx", "碳盤查績效統計"),
    (r"c:\Tandy資料夾\APPdata\supply-chain-carbon-management\scope3_ghg\outputs\SCOPE3_Report_GE115_各單位績效_20260513_1514.xlsx", "碳盤查績效統計（資分）")
]

for file_path, sheet_name in files:
    print(f"\n{'='*50}")
    print(f"File: {file_path}")
    print(f"Sheet: {sheet_name}")
    
    if not os.path.exists(file_path):
        print("Error: File not found.")
        continue

    # data_only=False
    wb_f = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    ws_f = wb_f[sheet_name]
    
    # data_only=True
    wb_t = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws_t = wb_t[sheet_name]
    
    print("\n--- Top 10 rows (A-C) [data_only=False] ---")
    for row in range(1, 11):
        vals = [ws_f.cell(row=row, column=col).value for col in range(1, 4)]
        print(f"Row {row:2}: {vals}")

    print("\n--- Target Row Analysis ---")
    found = False
    for row in range(1, ws_f.max_row + 1):
        a_val = ws_f.cell(row=row, column=1).value
        # Check if "2 供應商淨零承諾達成率" is in the cell
        if a_val and "2 供應商淨零承諾達成率" in str(a_val):
            found = True
            print(f"Found at Row: {row}")
            cols = ['A', 'B', 'C']
            for i, col_name in enumerate(cols, 1):
                raw = ws_f.cell(row=row, column=i).value
                calc = ws_t.cell(row=row, column=i).value
                print(f"Column {col_name}:")
                print(f"  data_only=False: {raw}")
                print(f"  data_only=True : {calc}")
            break
    if not found:
        print("Target row '2 供應商淨零承諾達成率' not found.")


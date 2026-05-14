from datetime import datetime

from openpyxl import Workbook

from scope3_ghg.config import DEPT_ORDER
from scope3_ghg.models import ContractRow
from scope3_ghg.reports.matrix_report import render_matrix_report
from scope3_ghg.reports.performance_report import MAIN_SHEET_NAME, render_performance_report


def build_row(
    contract_id: str,
    requester_unit: str,
    vendor_name_raw: str,
    net_zero_commitment_raw: str,
    contract_amount: float = 30_000_000,
) -> ContractRow:
    return ContractRow(
        contract_id=contract_id,
        contract_name=f"contract-{contract_id}",
        contract_amount=contract_amount,
        requester_unit=requester_unit,
        requester_name="tester",
        contract_status_raw="資料通過顧問審核標準",
        kpi_key="盤查範圍契約",
        vendor_name_raw=vendor_name_raw,
        net_zero_commitment_raw=net_zero_commitment_raw,
        iso14064_evidence_raw="",
    )


def test_report_renderers_create_expected_sheets_and_kpi_b_row() -> None:
    rows = [
        build_row("A-001", "政府應用處", "供應商甲", "未簽署"),
        build_row("A-002", "政府應用處", "供應商甲", "未簽署"),
        build_row("B-001", "營運系統管理處", "供應商乙", "已簽署"),
        build_row("C-001", "未知單位", "供應商丙", "已簽署"),
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)

    render_matrix_report(workbook, rows, "source.xlsx", now=datetime(2026, 5, 13, 15, 35))
    render_performance_report(workbook, rows, now=datetime(2026, 5, 13, 15, 35))

    expected_sheets = ["統計分析_矩陣", MAIN_SHEET_NAME, *[f"碳盤查績效統計_{dept}" for dept in DEPT_ORDER]]
    for sheet_name in expected_sheets:
        assert sheet_name in workbook.sheetnames

    main_ws = workbook[MAIN_SHEET_NAME]
    target_row = None
    for row in main_ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] == "2 供應商淨零承諾達成率":
            target_row = row
            break

    assert target_row is not None
    assert target_row[1] == 0.67
    assert target_row[2] == 3.33